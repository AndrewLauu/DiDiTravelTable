# -*- coding: utf-8 -*-
# from bs4 import BeautifulSoup as bs
import openpyxl as xl
from itertools import islice
from nodeTmp import etNode
import xml.etree.ElementTree as et
from lxml import etree as et
import time

# from nodeTmp import bsNode
print('Loaded libs.')

# 读取模板
workbook = xl.load_workbook('files/template.xlsx', read_only=True, data_only=True)
worksheet = workbook['tmp']
# worksheet = workbook.worksheets[0]
nRow = worksheet.max_row
workTable = worksheet.iter_rows(min_row=2, values_only=True)
# priceCol = worksheet.iter_rows(min_col=8, max_col=8, min_row=2, values_only=True)
# priceCol = list(zip(*priceCol))[0]
# sumPrice = sum([float(i) for i in priceCol])
sumPrice = workbook['价格']['a13'].value
# 判断行数，拆分页码（分表）
# pg0:0 -> 13r
# pg1:13 + 18 * (1-1) -> 13 + 18 * 1
# pg2:13+18 * (2-1) -> 13 + 18 * 2

nPage = 1+(nRow - 1 - 13) // 18 +1
table0 = islice(workTable, 0, 13)
tables = {p: islice(workTable, 13 + 18 * p - 18, 13 + 18 * p)
          for p in range(1, nPage)}
tables.update({0: table0})
tables = dict(sorted(tables.items(), key=lambda d: d[0]))

# 注册命名空间，实例化doc
# doc = bs(docTmp, 'xml')
for prefix, uri in etNode.xmlns.items():
    et.register_namespace(prefix, uri)
w = '{http://schemas.openxmlformats.org/wordprocessingml/2006/main}'
doc = et.fromstring(etNode.docTmp)

# 增加信息段
# @nRow
# @sumPrice
pInfo1 = et.fromstring(etNode.pInfoTmp1)
pInfoTmp2 = etNode.pInfoTmp2.replace('@nRow', str(nRow - 1)).replace('@sumPrice', str(sumPrice))

# pInfo = bs(pInfoTmp, 'xml')
pInfo2 = et.fromstring(pInfoTmp2)
pInfoSuffix = et.fromstring(etNode.pInfoTmpSuffix)
doc.find(w + 'body').append(pInfo1)
doc.find(w + 'body').append(pInfo2)
doc.find(w + 'body').append(pInfoSuffix)
print('Read excel.')

# 实例化padding, blankRow
# pPadding = bs(pPaddingTmp, 'xml')
# pBlank = bs(pBlankTmp, 'xml')
pPadding = et.fromstring(etNode.pPaddingTmp)
pBlank = et.fromstring(etNode.pBlankTmp)

# 建立偶行格式
# w_shd_0 = tr.new_tag('w:shd', attrs={'w:val' :"clear", 'w:color':"auto",'w:fill':"666666"})
# w_shd_even = tr.new_tag('w:shd', attrs={'w:val':  'clear', 'w:color': 'auto',
#                                         'w:fill': 'F0F0F0'})
# w_shd_even = et.SubElement(tr, 'w:shd', attrib={'w:val': 'clear', 'w:color': 'auto',
#               'w:fill': 'F0F0F0'})
w_shd_even = et.fromstring(etNode.w_shd_even_tmp)

# 遍历表中的分表，完善表元素并插入doc
for page, table in tables.items():
# 实例化表格
# tbl = bs(tblTmp, 'xml')
    tbl = et.fromstring(etNode.tblTmp)

    print(f'scanning table {page}')
# 遍历分表中的行，完善行元素并插入tbl
    for row in table:
# 实例化行元素
# tr = bs(trTmp, 'xml')
        tr = et.fromstring(etNode.trTmp)
# 遍历行中单元格，完善单元格元素，并插入tr
        for cell in row:
# 实例化单元格元素，判断奇偶行,匹配格式
# tc = bs(tcTmp, 'xml')
            tc = et.fromstring(etNode.tcTmp)
            print(f'Scanning table {page} row {row[0]} {cell}')
            if row[0] % 2 == 0:
                tc.find(w + 'tcPr').append(w_shd_even)
# w_t = tc.new_tag('w:t')
# w_t.string = str(cell)
            w_t = et.SubElement(tc.find(f'{w}p/{w}r'), w + 't')
            w_t.text = str(cell).replace('None','')
# tc.find(w + 'r').append(w_t)
            tr.append(tc)
# tr插入tbl
        tbl.append(tr)
# 完善tbl，插入doc
    print('Adding prefix')
    doc.find(w + 'body').append(pPadding)
    for _ in range(3):
        doc.find(w + 'body').append(pBlank)
    doc.find(w + 'body').append(tbl)
    time.sleep(1)

# 实例化sectPr并插入doc
# sectPr = bs(sectPrTmp, 'xml')
et.SubElement(doc.find(w + 'body'), w + 'p')
sectPr = et.fromstring(etNode.sectPrTmp)
doc.find(w + 'body').append(sectPr)
workbook.close()
print('write to file')
with open('files/DiDiTravelPersonnel/word/document.xml', 'w') as f:
    #f.write('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>')
    f.write(et.tostring(doc, encoding='utf-8',standalone=True).decode('utf-8'))
